import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from datetime import datetime
import re
import os
import warnings

# Suppress warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class GarmentsAutomationApp:
    def __init__(self, root):
        self.root = root
        self.root.title("IE Automation System (V34 - Working Day Fix)")
        self.root.geometry("1000x750")
        
        # Variables
        self.supervisor_path = tk.StringVar()
        self.master_path = tk.StringVar()
        
        self.create_menu()
        self.create_widgets()

    def create_menu(self):
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help & Support ❓", menu=help_menu)
        help_menu.add_command(label="🏢 Application Info", command=self.show_about_info)
        help_menu.add_separator()
        help_menu.add_command(label="❌ Exit", command=self.root.quit)

    def show_about_info(self):
        about_window = tk.Toplevel(self.root)
        about_window.title("Application Ownership & Support")
        about_window.geometry("550x650")
        about_window.resizable(False, False)
        about_window.configure(bg="#ecf0f1")

        header_frame = tk.Frame(about_window, bg="#2c3e50", pady=15)
        header_frame.pack(fill="x")
        tk.Label(header_frame, text="🏢 Application Ownership & Branding", 
                 font=("Arial", 13, "bold"), fg="#f1c40f", bg="#2c3e50").pack()

        factory_frame = tk.Frame(about_window, bg="#ecf0f1", padx=20, pady=10)
        factory_frame.pack(fill="x")
        tk.Label(factory_frame, text="This application is an in-house system developed for operational use at:", 
                 font=("Arial", 9, "italic"), bg="#ecf0f1", fg="#7f8c8d").pack(pady=(0, 5))
        tk.Label(factory_frame, text="Sonia and Sweaters Limited", 
                 font=("Arial", 14, "bold"), bg="#ecf0f1", fg="#2c3e50").pack()
        address_text = ("Factory Address:\n"
                        "Plot No: 604, Kondolbag, Taibpur\n"
                        "Dhaka–Ashulia Highway, Ashulia–1341, Bangladesh")
        tk.Label(factory_frame, text=address_text, justify="center",
                 font=("Arial", 10), bg="#ecf0f1", fg="#34495e").pack(pady=5)

        ttk.Separator(about_window, orient="horizontal").pack(fill="x", padx=20, pady=5)

        dev_header_frame = tk.Frame(about_window, bg="#ecf0f1", pady=5)
        dev_header_frame.pack()
        tk.Label(dev_header_frame, text="👨‍💻 Application Development & Support", 
                 font=("Arial", 12, "bold"), bg="#ecf0f1", fg="#2980b9").pack()

        dev_frame = tk.Frame(about_window, bg="#ecf0f1", padx=20, pady=5)
        dev_frame.pack(fill="x")
        tk.Label(dev_frame, text="Designed, Developed & Maintained by:", 
                 font=("Arial", 9), bg="#ecf0f1", fg="#7f8c8d").pack()
        tk.Label(dev_frame, text="Prottoy Saha", 
                 font=("Arial", 16, "bold"), bg="#ecf0f1", fg="#2c3e50").pack()
        tk.Label(dev_frame, text="Software Engineer (Internal Systems & Automation)", 
                 font=("Arial", 11, "bold"), bg="#ecf0f1", fg="#e67e22").pack()
        tk.Label(dev_frame, text="Sonia and Sweaters Limited", 
                 font=("Arial", 10), bg="#ecf0f1", fg="#34495e").pack(pady=(0, 10))

        contact_frame = tk.Frame(dev_frame, bg="#dfe6e9", padx=10, pady=10, relief="groove", bd=1)
        contact_frame.pack(pady=5)
        tk.Label(contact_frame, text="📞 Contact: +880 1745-547578", 
                 font=("Consolas", 11, "bold"), bg="#dfe6e9", fg="#27ae60").pack(anchor="w")
        tk.Label(contact_frame, text="📧 Email:   prottoy.saha@soniagroup.com", 
                 font=("Consolas", 11, "bold"), bg="#dfe6e9", fg="#2980b9").pack(anchor="w")

        notice_frame = tk.LabelFrame(about_window, text="⚠️ Support Notice", 
                                     font=("Arial", 9, "bold"), fg="#c0392b", bg="#ecf0f1", padx=10, pady=10)
        notice_frame.pack(fill="x", padx=20, pady=15)
        notice_msg = ("For any technical issues, system errors, or operational inconvenience,\n"
                      "please contact the above developer for support and assistance.")
        tk.Label(notice_frame, text=notice_msg, justify="center",
                 font=("Arial", 9), bg="#ecf0f1", fg="#7f8c8d").pack()
        tk.Button(about_window, text="Close", command=about_window.destroy, 
                  bg="#95a5a6", fg="white", width=15).pack(pady=10)

    def create_widgets(self):
        tk.Label(self.root, text="IE Automation System (Production & Efficiency)", font=("Arial", 16, "bold"), fg="#27ae60").pack(pady=15)
        main_frame = tk.Frame(self.root, padx=20)
        main_frame.pack(fill="both", expand=True)

        tk.Label(main_frame, text="1. Supervisor File (Daily Report):", font=("Arial", 10, "bold")).pack(anchor="w")
        tk.Entry(main_frame, textvariable=self.supervisor_path, width=80).pack(anchor="w", pady=5)
        tk.Button(main_frame, text="Browse Supervisor File", command=self.browse_sup, bg="#3498db", fg="white").pack(anchor="w", pady=(0, 15))

        tk.Label(main_frame, text="2. Master File (Linking Graph):", font=("Arial", 10, "bold")).pack(anchor="w")
        tk.Entry(main_frame, textvariable=self.master_path, width=80).pack(anchor="w", pady=5)
        tk.Button(main_frame, text="Browse Master File", command=self.browse_mas, bg="#3498db", fg="white").pack(anchor="w", pady=(0, 15))

        tk.Button(self.root, text="START AUTOMATION", command=self.run_process, font=("Arial", 12, "bold"), bg="#2c3e50", fg="white", height=2, width=30).pack(pady=10)

        self.progress = ttk.Progressbar(self.root, orient="horizontal", length=900, mode="determinate")
        self.progress.pack(pady=10)

        tk.Label(self.root, text="Process Log:", font=("Arial", 9, "bold")).pack(anchor="w", padx=20)
        self.log_text = scrolledtext.ScrolledText(self.root, height=18, width=110, font=("Consolas", 9))
        self.log_text.pack(padx=20, pady=10)
        tk.Label(self.root, text="© Sonia & Sweaters Limited | Internal Systems", font=("Arial", 8), fg="#95a5a6").pack(side="bottom", pady=5)

    def browse_sup(self):
        f = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xlsm")])
        if f: self.supervisor_path.set(f)

    def browse_mas(self):
        f = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xlsm")])
        if f: self.master_path.set(f)

    def log(self, msg):
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.log_text.insert(tk.END, f"{timestamp} - {msg}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def clean_text_strict(self, text):
        if not text: return ""
        return str(text).replace("\n", "").replace(".", "").replace(" ", "").strip().lower()

    def clean_header_loose(self, text):
        if not text: return ""
        return str(text).replace("\n", " ").strip().lower()

    def set_cell_value(self, cell, value):
        cell.value = value
        if cell.font:
            new_font = Font(name=cell.font.name, size=cell.font.size, 
                            bold=cell.font.bold, italic=cell.font.italic,
                            color="000000") 
            cell.font = new_font

    def run_process(self):
        sup_file = self.supervisor_path.get()
        mas_file = self.master_path.get()

        if not sup_file or not mas_file:
            messagebox.showerror("Error", "Please select both files.")
            return

        try:
            self.log("--- STARTED ---")
            self.progress['value'] = 5
            
            self.log(f"Step 1: Reading Supervisor File...")
            extracted_data = self.read_supervisor_file(sup_file)
            
            if not extracted_data:
                self.log("CRITICAL: No valid production data found.")
                messagebox.showwarning("No Data", "No data found where 'Today' > 0.")
                return
            
            self.log(f"Collected {len(extracted_data)} valid data entries.")
            self.progress['value'] = 40
            
            self.log(f"Step 2: Updating Master File with Formulas...")
            updated_count = self.update_master_file(mas_file, extracted_data)
            
            self.progress['value'] = 100
            self.log(f"--- COMPLETED: Updated {updated_count} Sheets ---")
            messagebox.showinfo("Success", f"Done! Updated {updated_count} sheets in the Master File.")

        except Exception as e:
            self.log(f"CRITICAL ERROR: {str(e)}")
            messagebox.showerror("Error", str(e))

    # ==========================================
    # STEP 1: READ SUPERVISOR
    # ==========================================
    def find_date_in_row(self, row):
        date_pattern = re.compile(r'(\d{1,2})[-./](\d{1,2})[-./](\d{2,4})')
        for cell in row:
            if not cell: continue
            if isinstance(cell, datetime): return cell
            s_val = str(cell).strip()
            match = date_pattern.search(s_val)
            if match:
                try:
                    d_str = f"{match.group(1)}/{match.group(2)}/{match.group(3)}"
                    for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"]:
                        try: return datetime.strptime(d_str, fmt)
                        except: continue
                except: pass
        return None

    def read_supervisor_file(self, filepath):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        all_data = []

        for sheet in wb.worksheets:
            self.log(f"Scanning sheet: {sheet.title}")
            rows = list(sheet.iter_rows(values_only=True))
            
            row_date_map = {}
            for r_idx, row in enumerate(rows):
                found_date = self.find_date_in_row(row)
                if found_date: row_date_map[r_idx] = found_date
            
            for r_idx, row in enumerate(rows):
                if not row: continue
                row_strict = [self.clean_text_strict(c) for c in row]
                row_loose = [self.clean_header_loose(c) for c in row]
                
                if "styleno" in row_strict and "today" in row_strict:
                    
                    closest_date = None
                    for search_back in range(r_idx, max(-1, r_idx-50), -1):
                        if search_back in row_date_map:
                            closest_date = row_date_map[search_back]
                            break
                    if not closest_date: closest_date = self.find_date_in_row(row)
                    if not closest_date: continue

                    col = {}
                    for idx, val in enumerate(row_loose):
                        v_strict = row_strict[idx]
                        if "style no" in val: col['style'] = idx
                        elif "today" in val: col['today'] = idx 
                        elif "buyer" in val: col['buyer'] = idx
                        elif "gg" == v_strict: col['gg'] = idx
                        elif "order" in val and "qty" in val: col['order_qty'] = idx
                        elif "con" in val and "qty" in val: col['con_qty'] = idx
                        elif val == "m/c" or val == "mc": col['mc'] = idx
                        elif "working" in val and "min" in val: col['min'] = idx
                        elif "smv" in val: col['smv'] = idx
                        elif "aver" in val and "min" in val: col['aver_min'] = idx

                    j = r_idx + 1
                    while j < len(rows):
                        d_row = rows[j]
                        j += 1
                        if not any(d_row): break
                        if d_row[0] and "total" in str(d_row[0]).lower(): break
                        if j-1 in row_date_map and j-1 > r_idx + 1: break 

                        try:
                            s_name = d_row[col['style']]
                            prod = d_row[col['today']]
                            
                            valid = False
                            if prod is not None:
                                try:
                                    if float(prod) > 0: valid = True
                                except: pass
                            
                            if valid and s_name:
                                mc_val = d_row[col['mc']] if 'mc' in col else 0
                                min_val = d_row[col['min']] if 'min' in col else 0
                                smv_val = d_row[col['smv']] if 'smv' in col else 0
                                aver_min = 0
                                if 'aver_min' in col: aver_min = d_row[col['aver_min']]
                                elif 'min' in col: aver_min = d_row[col['min']]
                                
                                try: mc = float(mc_val)
                                except: mc = 0
                                try: mins = float(min_val)
                                except: mins = 0
                                try: out = float(prod)
                                except: out = 0
                                try: smv = float(smv_val)
                                except: smv = 0
                                try: a_min = float(aver_min)
                                except: a_min = 0

                                entry = {
                                    "date": closest_date,
                                    "style": str(s_name).strip(),
                                    "output": prod,
                                    "buyer": d_row[col.get('buyer')] if 'buyer' in col else "",
                                    "gg": d_row[col.get('gg')] if 'gg' in col else "",
                                    "order_qty": d_row[col.get('order_qty')] if 'order_qty' in col else "",
                                    "con_qty": d_row[col.get('con_qty')] if 'con_qty' in col else "",
                                    "mc": mc,
                                    "min": mins,
                                    "aver_min": a_min,
                                    "smv": smv
                                }
                                all_data.append(entry)
                                self.log(f" -> Found: {s_name} | Qty: {prod}")
                        except: continue
        return all_data

    # ==========================================
    # STEP 2: WRITE MASTER
    # ==========================================
    def update_master_file(self, filepath, data):
        wb = openpyxl.load_workbook(filepath)
        if "FORMATE" not in wb.sheetnames:
            self.log("ERROR: 'FORMATE' sheet missing!")
            return 0

        fmt = wb["FORMATE"]
        updated_counter = 0
        data.sort(key=lambda x: x['date'])

        for entry in data:
            raw_style = str(entry['style']).strip()
            clean_style = re.sub(r'[\s\-_]*REQ$', '', raw_style, flags=re.IGNORECASE).strip()
            entry['style'] = clean_style
            safe_style = re.sub(r'[\\/*?:\[\]]', '_', clean_style)[:31]
            
            ws = None
            if safe_style in wb.sheetnames:
                ws = wb[safe_style]
            else:
                ws = wb.copy_worksheet(fmt)
                ws.title = safe_style
                self.force_fill_headers(ws, entry)
                self.log(f"Created new sheet: {safe_style}")

            self.force_fill_headers(ws, entry)
            if entry['smv']: self.set_cell_value(ws['K11'], entry['smv'])

            header_row_idx = self.find_table_header(ws)
            if not header_row_idx: continue

            col_map = self.map_table_columns(ws, header_row_idx)
            
            target_row = None
            r = header_row_idx + 1
            last_day_val = 0
            
            while r < 1000:
                c_date = ws.cell(r, col_map['Date']).value
                c_day = ws.cell(r, col_map.get('Day', 1)).value
                if isinstance(c_day, int): last_day_val = c_day
                
                is_match = False
                if isinstance(c_date, datetime) and c_date.date() == entry['date'].date(): is_match = True
                elif isinstance(c_date, str) and entry['date'].strftime("%d-%b") in str(c_date): is_match = True
                
                if is_match:
                    target_row = r
                    break
                if c_date is None:
                    target_row = r
                    break
                r += 1
            
            if target_row:
                if ws.cell(target_row, col_map['Date']).value is None:
                    if 'Day' in col_map: 
                        self.set_cell_value(ws.cell(target_row, col_map['Day']), last_day_val + 1)
                    self.set_cell_value(ws.cell(target_row, col_map['Date']), entry['date'].strftime("%d-%b"))
                
                self.set_cell_value(ws.cell(target_row, col_map['Output']), entry['output'])
                if 'MC' in col_map: 
                    self.set_cell_value(ws.cell(target_row, col_map['MC']), entry['mc'])

                c_op = get_column_letter(col_map['MC']) if 'MC' in col_map else 'C'
                c_out = get_column_letter(col_map['Output'])
                c_min = get_column_letter(col_map['Min']) if 'Min' in col_map else 'F'
                row = target_row
                aver_min = entry['aver_min']

                if 'AvgProd' in col_map:
                    f_avg = f"={c_out}{row}/{c_op}{row}"
                    self.set_cell_value(ws.cell(row, col_map['AvgProd']), f_avg)

                if 'Min' in col_map:
                    f_min = f"={c_op}{row}*{aver_min}"
                    self.set_cell_value(ws.cell(row, col_map['Min']), f_min)

                if 'Eff' in col_map:
                    f_eff = f"={c_out}{row}*$K$11/{c_min}{row}"
                    cell_eff = ws.cell(row, col_map['Eff'])
                    cell_eff.value = f_eff
                    cell_eff.number_format = '0%' 
                    self.set_cell_value(cell_eff, f_eff)

                if 'Time' in col_map:
                    f_time = f"=({c_min}{row}/{c_op}{row})/60"
                    self.set_cell_value(ws.cell(row, col_map['Time']), f_time)
                
                updated_counter += 1
                
                self.update_footer_formulas(ws, header_row_idx, col_map)
                if 'Eff' in col_map and 'Day' in col_map:
                    self.add_efficiency_chart(ws, header_row_idx, col_map)
            
            # UPDATE TOTAL SUMMARY
            self.update_total_summary(wb, ws, safe_style, header_row_idx, col_map)

        wb.save(filepath)
        return updated_counter

    def update_total_summary(self, wb, ws_style, style_name, header_row, col_map):
        if "TOTAL SUMMARY" not in wb.sheetnames: return
        ws_sum = wb["TOTAL SUMMARY"]
        
        addr_map = {}
        total_row = None
        for r in range(header_row + 1, 100):
            val = ws_style.cell(r, 1).value
            if val and ("total" in str(val).lower()):
                total_row = r
                break
        
        for r in range(1, 70):
            for c in range(1, 15):
                val = ws_style.cell(r, c).value
                if not val: continue
                v = self.clean_text_strict(val)
                if "customer" in v or "style" in v or "gauge" in v or "orderqty" in v or "linkingqty" in v or "smv" in v:
                    end_col = c
                    for merged_range in ws_style.merged_cells.ranges:
                        if ws_style.cell(r, c).coordinate in merged_range:
                            end_col = merged_range.max_col
                            break
                    target_c = end_col + 1
                    target_cell = ws_style.cell(r, target_c)
                    final_target = target_cell
                    for merged_range in ws_style.merged_cells.ranges:
                        if target_cell.coordinate in merged_range:
                            final_target = ws_style.cell(merged_range.min_row, merged_range.min_col)
                            break
                    
                    if "customer" in v: addr_map['Buyer'] = final_target.coordinate
                    elif "style" in v: addr_map['Style'] = final_target.coordinate
                    elif "gauge" in v: addr_map['GG'] = final_target.coordinate
                    elif "smv" in v: addr_map['SMV'] = final_target.coordinate
                    elif "orderqty" in v: addr_map['OrderQty'] = final_target.coordinate
                    elif "linkingqty" in v: addr_map['LinkQty'] = final_target.coordinate

        target_r = None
        start_row = 5
        last_sno = 0
        for r in range(start_row, 500):
            s_val = ws_sum.cell(r, 3).value
            s_no = ws_sum.cell(r, 1).value
            if isinstance(s_no, int): last_sno = s_no
            if s_val and isinstance(s_val, str) and style_name in s_val:
                target_r = r
                break
            if not s_val:
                target_r = r
                break
        
        if not target_r: return

        if ws_sum.cell(target_r, 1).value is None:
            self.set_cell_value(ws_sum.cell(target_r, 1), last_sno + 1)

        if 'Buyer' in addr_map:
             val = ws_style[addr_map['Buyer']].value
             self.set_cell_value(ws_sum.cell(target_r, 2), val)
        
        self.set_cell_value(ws_sum.cell(target_r, 3), style_name)

        if 'GG' in addr_map:
             val = ws_style[addr_map['GG']].value
             self.set_cell_value(ws_sum.cell(target_r, 4), val)

        quoted_style = f"'{style_name}'"
        
        if 'OrderQty' in addr_map:
            f = f"={quoted_style}!{addr_map['OrderQty']}"
            self.set_cell_value(ws_sum.cell(target_r, 5), f)
            
        if 'LinkQty' in addr_map:
            f = f"={quoted_style}!{addr_map['LinkQty']}"
            self.set_cell_value(ws_sum.cell(target_r, 6), f)
            
        f_bal = f"=E{target_r}-F{target_r}"
        self.set_cell_value(ws_sum.cell(target_r, 7), f_bal)

        if 'SMV' in addr_map:
            f = f"={quoted_style}!{addr_map['SMV']}"
            self.set_cell_value(ws_sum.cell(target_r, 8), f)

        if total_row:
            col_out = get_column_letter(col_map['Output'])
            col_wmin = get_column_letter(col_map['Min']) if 'Min' in col_map else 'F'
            col_eff = get_column_letter(col_map['Eff']) if 'Eff' in col_map else 'G'
            
            f_out = f"={quoted_style}!{col_out}{total_row}"
            self.set_cell_value(ws_sum.cell(target_r, 9), f_out)
            
            f_wmin = f"={quoted_style}!{col_wmin}{total_row}"
            self.set_cell_value(ws_sum.cell(target_r, 10), f_wmin)
            
            f_pmin = f"=H{target_r}*I{target_r}"
            self.set_cell_value(ws_sum.cell(target_r, 11), f_pmin)
            
            # --- V34: FIXED COUNTA FORMULA ---
            col_date = get_column_letter(col_map['Date'])
            s_row = header_row + 1
            e_row = total_row - 1
            # Use COUNTA to count text dates (Non-Empty Cells)
            f_wday = f"=COUNTA({quoted_style}!{col_date}{s_row}:{col_date}{e_row})"
            self.set_cell_value(ws_sum.cell(target_r, 12), f_wday)
            # ---------------------------------
            
            f_eff = f"={quoted_style}!{col_eff}{total_row}"
            cell = ws_sum.cell(target_r, 13)
            cell.value = f_eff
            cell.number_format = '0%'
            self.set_cell_value(cell, f_eff)

    def update_footer_formulas(self, ws, header_row, col_map):
        total_row = None
        for r in range(header_row + 1, 100):
            val = ws.cell(r, 1).value
            if val and ("total" in str(val).lower()):
                total_row = r
                break
        
        if total_row:
            start = header_row + 1
            end = total_row - 1
            c_op = get_column_letter(col_map['MC']) if 'MC' in col_map else None
            c_out = get_column_letter(col_map['Output'])
            c_avg = get_column_letter(col_map['AvgProd']) if 'AvgProd' in col_map else None
            c_min = get_column_letter(col_map['Min']) if 'Min' in col_map else None
            c_eff = get_column_letter(col_map['Eff']) if 'Eff' in col_map else None
            
            if c_op: self.set_cell_value(ws.cell(total_row, col_map['MC']), f"=SUM({c_op}{start}:{c_op}{end})")
            self.set_cell_value(ws.cell(total_row, col_map['Output']), f"=SUM({c_out}{start}:{c_out}{end})")
            if c_avg and c_op: self.set_cell_value(ws.cell(total_row, col_map['AvgProd']), f"={c_out}{total_row}/{c_op}{total_row}")
            if c_min: self.set_cell_value(ws.cell(total_row, col_map['Min']), f"=SUM({c_min}{start}:{c_min}{end})")
            if c_eff and c_min:
                f_footer_eff = f"={c_out}{total_row}*$K$11/{c_min}{total_row}"
                cell = ws.cell(total_row, col_map['Eff'])
                cell.value = f_footer_eff
                cell.number_format = '0%'
                self.set_cell_value(cell, f_footer_eff)

    def add_efficiency_chart(self, ws, header_row, col_map):
        if len(ws._charts) > 0: del ws._charts[:]
        
        eff_col = col_map['Eff']
        day_col = col_map['Day']
        data_start = header_row + 1
        data_end = header_row + 32 
        
        values = Reference(ws, min_col=eff_col, min_row=header_row, max_row=data_end)
        cats = Reference(ws, min_col=day_col, min_row=data_start, max_row=data_end)
        
        chart = LineChart()
        chart.title = "Efficiency (%)"
        chart.style = 13
        chart.y_axis.title = "Efficiency"
        chart.x_axis.title = "Day"
        chart.legend = None 
        
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)
        
        s1 = chart.series[0]
        s1.marker.symbol = "circle"
        s1.dLbls = DataLabelList() 
        s1.dLbls.showVal = True
        s1.dLbls.numFmt = '0%'
        
        chart.height = 10 
        chart.width = 18  
        ws.add_chart(chart, "J16")

    def force_fill_headers(self, ws, entry):
        fill_map = {
            "style": entry['style'],
            "customer": entry['buyer'],
            "gauge": entry['gg'],
            "orderqty": entry['order_qty'],
            "consumtionqty": entry['con_qty']
        }

        for r in range(1, 65):
            for c in range(1, 15):
                cell = ws.cell(r, c)
                if not cell.value: continue
                val = self.clean_text_strict(cell.value)
                
                if val in fill_map:
                    end_col = c
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            end_col = merged_range.max_col
                            break
                    target_c = end_col + 1
                    target_cell = ws.cell(r, target_c)
                    final_target = target_cell
                    for merged_range in ws.merged_cells.ranges:
                        if target_cell.coordinate in merged_range:
                            final_target = ws.cell(merged_range.min_row, merged_range.min_col)
                            break
                    self.set_cell_value(final_target, fill_map[val])

    def find_table_header(self, ws):
        for r in range(10, 40):
            row_vals = []
            for c in range(1, 20):
                val = ws.cell(r, c).value
                row_vals.append(self.clean_text_strict(val) if val else "")
            if ("output" in row_vals or "production" in row_vals) and \
               ("date" in row_vals or any("date" in x for x in row_vals)):
                return r
        return None

    def map_table_columns(self, ws, r):
        m = {}
        for c in range(1, 20):
            val = ws.cell(r, c).value
            if not val: continue
            v = str(val).lower()
            
            if "day" in v and "days" not in v: m['Day'] = c
            elif "date" in v: m['Date'] = c
            elif "output" in v: m['Output'] = c
            
            elif "op" in v and ("no" in v or "m/c" in v): m['MC'] = c
            elif "avg" in v and "prod" in v: m['AvgProd'] = c
            elif "total" in v and ("work" in v or "min" in v): m['Min'] = c
            elif "eff" in v: m['Eff'] = c
            elif "time" in v: m['Time'] = c
        return m

if __name__ == "__main__":
    root = tk.Tk()
    app = GarmentsAutomationApp(root)
    root.mainloop()