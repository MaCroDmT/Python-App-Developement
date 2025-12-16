#-----Developer--------#
#-----MaCroDmT--------#
#-----Owner-----------#
#-----Prottoy-Saha----#
#----prottoys28@gmail.com---#
#----Bangladeshi----------#
#----AIUB'ian------------#
#----21-3--------------#

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import threading
import time
import os
import pandas as pd
import json
from datetime import datetime, date, timedelta
import ctypes
import pyautogui
import shutil
import urllib.request
import sys

# --- Restored Imports for Excel/Zip Features ---
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import io
import zipfile

# --- STABLE PDF GENERATION IMPORT ---
# Replaced docx2pdf with reportlab for stability in EXEs
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from pypdf import PdfWriter, PdfReader
except ImportError:
    print("Please install reportlab: pip install reportlab pypdf")
    pass

# ==========================================
# UNIFIED MONITOR APPLICATION
# ==========================================
class UnifiedMonitorApp:
    def __init__(self, master, username, data_dir):
        self.master = master
        self.username = username
        self.data_dir = data_dir
        
        # --- ADMIN CONTROL SETTINGS ---
        self.ADMIN_URL = "https://gist.githubusercontent.com/raw/PLACEHOLDER_FOR_YOUR_ADMIN_FILE.txt" 
        self.admin_check_interval = 60 
        # -----------------------------

        self.is_working = False 
        self.on_break = False 
        
        self.log_data = [] 
        self.start_time = None
        self.break_start_time = None
        
        self.report_password = "zia@SSL"
        self.has_warned_overtime = False

        self.activity_excel_path = None
        self.summary_excel_path = None

        self.screenshot_dir = os.path.join(self.data_dir, "screenshots")
        os.makedirs(self.screenshot_dir, exist_ok=True)
        
        self.daily_backup_file = os.path.join(self.data_dir, f"backup_log_{date.today()}.json")
        
        self.screenshot_thread = None
        self.activity_thread = None
        self.admin_thread = None

        self.cleanup_old_data()
        self.load_daily_backup()
        
        if not self.check_remote_permission(initial=True):
            sys.exit()

        self.setup_ui()
        self.update_live_timer()
        
        self.start_admin_listener()
        
        self.master.protocol("WM_DELETE_WINDOW", self.on_closing)

    # ... [Admin Control Logic - Unchanged] ...
    def check_remote_permission(self, initial=False):
        if "PLACEHOLDER" in self.ADMIN_URL: return True 
        try:
            with urllib.request.urlopen(self.ADMIN_URL, timeout=5) as response:
                status = response.read().decode('utf-8').strip().upper()
            if "STOP" in status:
                if initial: messagebox.showerror("Access Denied", "Administrator has disabled this application.")
                else: self.master.after(0, self.force_remote_shutdown)
                return False
            return True
        except: return True

    def start_admin_listener(self):
        self.admin_thread = threading.Thread(target=self.admin_loop, daemon=True)
        self.admin_thread.start()

    def admin_loop(self):
        while True:
            time.sleep(self.admin_check_interval)
            self.check_remote_permission(initial=False)

    def force_remote_shutdown(self):
        if self.is_working: self.clock_out(silent=True) 
        messagebox.showwarning("Admin Alert", "The Administrator has remotely disabled this session.\nThe application will now close.")
        self.master.destroy()
        sys.exit()

    # ... [Cleanup & UI Logic - Unchanged] ...
    def cleanup_old_data(self):
        today_str = date.today().strftime("%Y%m%d")
        for filename in os.listdir(self.screenshot_dir):
            if filename.startswith("screenshot_") and filename.endswith(".png"):
                try:
                    if filename.split("_")[1] != today_str:
                        os.remove(os.path.join(self.screenshot_dir, filename))
                except: pass
        for filename in os.listdir(self.data_dir):
            if filename.startswith("backup_log_") and filename.endswith(".json"):
                if filename != f"backup_log_{date.today()}.json":
                    os.remove(os.path.join(self.data_dir, filename))

    def setup_ui(self):
        main_frame = tk.Frame(self.master, padx=20, pady=20)
        main_frame.pack(expand=True, fill="both")

        self.greeting_label = tk.Label(main_frame, text=f"Welcome, {self.username}", font=("Arial", 14))
        self.greeting_label.pack(pady=(0, 5))

        self.time_label = tk.Label(main_frame, text="", font=("Arial", 10), fg="gray")
        self.time_label.pack(pady=5)
        self.update_time_display()

        self.live_timer_label = tk.Label(main_frame, text="Today's Work: 00:00:00", font=("Arial", 18, "bold"), fg="#0056b3")
        self.live_timer_label.pack(pady=15)

        self.status_label = tk.Label(main_frame, text="Status: Ready", font=("Arial", 12, "bold"))
        self.status_label.pack(pady=10)

        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(pady=20)

        self.clock_in_btn = tk.Button(btn_frame, text="Clock-in", font=("Arial", 11, "bold"), width=15, bg="#4CAF50", fg="white", command=self.clock_in)
        self.clock_in_btn.grid(row=0, column=0, padx=5, pady=5)
        
        self.break_btn = tk.Button(btn_frame, text="Take a Break", font=("Arial", 11, "bold"), width=15, bg="#FFC107", fg="black", command=self.take_a_break)
        self.break_btn.grid(row=1, column=0, padx=5, pady=5)

        self.clock_out_btn = tk.Button(btn_frame, text="Clock-out", font=("Arial", 11, "bold"), width=15, bg="#F44336", fg="white", command=self.clock_out)
        self.clock_out_btn.grid(row=2, column=0, padx=5, pady=5)

        lbl_note = tk.Label(main_frame, text="* Controls Work Log AND Screen Monitor", font=("Arial", 8, "italic"))
        lbl_note.pack(pady=5)

        if self.log_data:
            tk.Label(main_frame, text=f"Restored {len(self.log_data)} activities.", fg="blue", font=("Arial", 9)).pack()

        contact_frame = tk.Frame(main_frame, pady=10)
        contact_frame.pack(side="bottom", fill="x", pady=(20, 0))
        tk.Label(contact_frame, text="For any types of inconvenience with this Application please contact with:", font=("Arial", 9, "bold"), fg="black").pack(anchor="center")
        tk.Label(contact_frame, text="Prottoy Saha", font=("Arial", 10, "bold"), fg="black").pack(anchor="center")
        tk.Label(contact_frame, text="Automation Engineer", font=("Arial", 9, "bold"), fg="black").pack(anchor="center")
        tk.Label(contact_frame, text="Sonia and Sweaters Limited", font=("Arial", 9, "bold"), fg="black").pack(anchor="center")
        tk.Label(contact_frame, text="+8801745547578 📞", font=("Arial", 9), fg="black").pack(anchor="center")
        tk.Label(contact_frame, text="prottoy.saha@soniagroup.com 📧", font=("Arial", 9), fg="black").pack(anchor="center")

    def clock_in(self):
        if self.is_working and not self.on_break:
            messagebox.showinfo("Info", "You are already working.")
            return
        if self.on_break:
            break_end = datetime.now()
            duration = break_end - self.break_start_time
            self.log_activity("Took a break from work", self.break_start_time, break_end, duration)
            self.on_break = False
            messagebox.showinfo("Resuming", "Break ended. Monitoring resumed.")
        else:
            self.is_working = True
            self.start_time = datetime.now()
            self.log_activity("Clocked in to work", self.start_time, self.start_time)
            messagebox.showinfo("Clocked In", "Work started. Monitoring active.")
        self.status_label.config(text="Status: Working...")
        self.start_threads_if_needed()

    def take_a_break(self):
        if not self.is_working or self.on_break:
            messagebox.showwarning("Warning", "You must be working to take a break.")
            return
        end_time = datetime.now()
        duration = end_time - self.start_time
        self.log_activity("Worked", self.start_time, end_time, duration)
        self.on_break = True
        self.break_start_time = datetime.now()
        self.status_label.config(text="Status: On Break (Paused)")
        messagebox.showinfo("On Break", "Enjoy your break!\n\nPlease click 'Clock-in' when you are finished to resume work.")

    def clock_out(self, silent=False):
        if not self.is_working:
            if not silent: messagebox.showwarning("Warning", "You must be clocked in to clock out.")
            return
        if self.on_break:
            duration = datetime.now() - self.break_start_time
            self.log_activity("Took a break from work", self.break_start_time, datetime.now(), duration)
        else:
            duration = datetime.now() - self.start_time
            self.log_activity("Worked", self.start_time, datetime.now(), duration)

        self.is_working = False
        self.on_break = False
        self.log_activity("Clocked out from work", datetime.now(), datetime.now())
        
        self.status_label.config(text="Status: Clocked Out")
        
        self.save_data_to_excel()
        self.save_summary_to_excel()
        if self.activity_excel_path and self.summary_excel_path:
            self.create_zip_file()

        # REPLACED OLD PDF FUNCTION WITH NEW STABLE ONE
        self.create_unified_report_stable() 

    def on_closing(self):
        if self.is_working or self.on_break:
            if messagebox.askokcancel("Quit", "Work session is active! Do you want to Clock Out and Save Reports before quitting?"):
                self.clock_out(silent=True)
                self.master.destroy()
        else:
            self.master.destroy()

    def start_threads_if_needed(self):
        if self.activity_thread is None or not self.activity_thread.is_alive():
            self.activity_thread = threading.Thread(target=self.monitor_activity_loop, daemon=True)
            self.activity_thread.start()
        if self.screenshot_thread is None or not self.screenshot_thread.is_alive():
            self.screenshot_thread = threading.Thread(target=self.screenshot_loop, daemon=True)
            self.screenshot_thread.start()

    def screenshot_loop(self):
        while self.is_working:
            if not self.on_break:
                try:
                    now = datetime.now()
                    filename = f"screenshot_{now.strftime('%Y%m%d_%H%M%S')}.png"
                    filepath = os.path.join(self.screenshot_dir, filename)
                    pyautogui.screenshot().save(filepath)
                except: pass
                for _ in range(300):
                    if not self.is_working: return
                    time.sleep(1)
            else: time.sleep(1)

    def monitor_activity_loop(self):
        last_net = self.check_internet()
        last_lock = self.check_lock()
        net_start = None
        lock_start = None
        while self.is_working:
            if self.on_break:
                time.sleep(1)
                continue
            time.sleep(1)
            curr_net = self.check_internet()
            if not curr_net and last_net: net_start = datetime.now()
            elif curr_net and not last_net and net_start:
                self.log_activity("Internet interrupted", net_start, datetime.now(), datetime.now() - net_start)
                net_start = None
            last_net = curr_net
            curr_lock = self.check_lock()
            if curr_lock and not last_lock: lock_start = datetime.now()
            elif not curr_lock and last_lock and lock_start:
                self.log_activity("System locked", lock_start, datetime.now(), datetime.now() - lock_start)
                lock_start = None
            last_lock = curr_lock

    def log_activity(self, act_type, start, end, duration=None):
        if duration:
            sec = int(duration.total_seconds())
            h, r = divmod(sec, 3600)
            m, s = divmod(r, 60)
            dur_str = f"{act_type} for {h:02d}h {m:02d}m {s:02d}s"
        else: dur_str = act_type
        
        entry = {
            "Employee Name": self.username,
            "Date": start.strftime("%Y-%m-%d"),
            "Start Time": start.strftime("%I:%M %p"),
            "End Time": end.strftime("%I:%M %p") if end else "-",
            "Activity Duration": dur_str
        }
        self.log_data.append(entry)
        self.save_backup()

    def update_live_timer(self):
        total_sec = 0
        for entry in self.log_data:
            if "Worked for" in entry['Activity Duration']:
                total_sec += self.parse_duration(entry['Activity Duration']).total_seconds()
        if self.is_working and not self.on_break:
            total_sec += (datetime.now() - self.start_time).total_seconds()
        h, r = divmod(int(total_sec), 3600)
        m, s = divmod(r, 60)
        self.live_timer_label.config(text=f"Today's Work: {h:02d}:{m:02d}:{s:02d}")
        if total_sec > 28800 and not self.has_warned_overtime:
            self.has_warned_overtime = True
            messagebox.showwarning("Balance Alert", "You've worked over 8 hours today!")
        self.master.after(1000, self.update_live_timer)

    def update_time_display(self):
        now = datetime.now()
        self.time_label.config(text=now.strftime("%A, %B %d, %Y | %I:%M:%S %p"))
        self.master.after(1000, self.update_time_display)

    def analyze_data(self, df):
        df['DurSec'] = df['Activity Duration'].apply(lambda x: self.parse_duration(x).total_seconds())
        work_sec = df[df['Activity Duration'].str.contains("Worked for", na=False)]['DurSec'].sum()
        break_sec = df[df['Activity Duration'].str.contains("Took a break", na=False)]['DurSec'].sum()
        locked_sec = df[df['Activity Duration'].str.contains("System locked", na=False)]['DurSec'].sum()
        net_sec = df[df['Activity Duration'].str.contains("Internet interrupted", na=False)]['DurSec'].sum()
        return {
            'Worked for in a day': str(timedelta(seconds=int(work_sec))),
            'System locked/clocked out': df[df['Activity Duration'].str.contains("System locked|Clocked out", na=False)].shape[0],
            'Took a break': df[df['Activity Duration'].str.contains("Took a break", na=False)].shape[0],
            'Internet Interrupted': df[df['Activity Duration'].str.contains("Internet interrupted", na=False)].shape[0],
            'Average Idle Time': str(timedelta(seconds=int((break_sec + locked_sec + net_sec)/max(1, (df.shape[0]))))), 
            'total_work_hours': work_sec / 3600,
            'total_break_hours': break_sec / 3600,
            'total_locked_hours': locked_sec / 3600,
            'total_interrupted_hours': net_sec / 3600
        }

    def save_data_to_excel(self):
        if not self.log_data: return
        try:
            df = pd.DataFrame(self.log_data)
            if 'Employee Name' not in df.columns: df.insert(0, 'Employee Name', self.username)
            else: df['Employee Name'] = df['Employee Name'].fillna(self.username)
            if 'Date' not in df.columns: df.insert(1, 'Date', date.today().strftime("%Y-%m-%d"))
            else: df['Date'] = df['Date'].fillna(date.today().strftime("%Y-%m-%d"))
            desired_order = ['Employee Name', 'Date', 'Start Time', 'End Time', 'Activity Duration']
            df = df.reindex(columns=desired_order)
            filename = f"{self.username.replace(' ', '')}-Work-log-Activity-{date.today().strftime('%Y-%m-%d')}.xlsx"
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            self.activity_excel_path = os.path.join(downloads_path, filename)
            wb = Workbook()
            ws = wb.active
            ws.title = "Activity Log"
            for r in dataframe_to_rows(df, index=False, header=True): ws.append(r)
            wb.save(self.activity_excel_path)
        except Exception as e:
            messagebox.showerror("Error", f"Could not save Activity Excel: {e}")
            self.activity_excel_path = None

    def save_summary_to_excel(self):
        if not self.log_data: return
        try:
            df = pd.DataFrame(self.log_data)
            analysis = self.analyze_data(df)
            summary_data = {
                'Metric': ['Worked for in a day', 'System locked/clocked out', 'Took a break', 'Internet Interrupted', 'Average Idle Time'],
                'Value': [analysis['Worked for in a day'], analysis['System locked/clocked out'], analysis['Took a break'], analysis['Internet Interrupted'], analysis['Average Idle Time']]
            }
            summary_df = pd.DataFrame(summary_data)
            chart_data = {'Metric': ['Work', 'Break', 'System Locked', 'Internet Interrupted'], 'Duration': [analysis['total_work_hours'], analysis['total_break_hours'], analysis['total_locked_hours'], analysis['total_interrupted_hours']]}
            wb = Workbook()
            ws = wb.active
            ws.title = "Summary Report"
            ws['A1'] = self.username
            ws['A3'] = datetime.now().strftime("%Y-%m-%d")
            for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), start=5):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 5: cell.font = Font(bold=True)
            fig, ax = plt.subplots(figsize=(8, 4))
            ax.barh(chart_data['Metric'], chart_data['Duration'])
            ax.set_title('Activity Breakdown')
            ax.set_xlabel('Duration (Hours)')
            plt.tight_layout()
            buf = io.BytesIO()
            fig.savefig(buf, format='png')
            buf.seek(0)
            img = OpenpyxlImage(buf)
            img.anchor = 'A15'
            ws.add_image(img)
            filename = f"{self.username.replace(' ', '')}-Work-Summary-Report-{date.today().strftime('%Y-%m-%d')}.xlsx"
            self.summary_excel_path = os.path.join(os.path.expanduser("~"), "Downloads", filename)
            wb.save(self.summary_excel_path)
            plt.close(fig)
        except Exception as e:
            messagebox.showerror("Error", f"Could not save Summary Excel: {e}")
            self.summary_excel_path = None

    def create_zip_file(self):
        try:
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            zip_filename = f"{self.username.replace(' ', '')}-Work-Reports-{date.today().strftime('%Y-%m-%d')}.zip"
            zip_file_path = os.path.join(downloads_path, zip_filename)
            with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                if self.activity_excel_path and os.path.exists(self.activity_excel_path): zf.write(self.activity_excel_path, os.path.basename(self.activity_excel_path))
                if self.summary_excel_path and os.path.exists(self.summary_excel_path): zf.write(self.summary_excel_path, os.path.basename(self.summary_excel_path))
            if self.activity_excel_path and os.path.exists(self.activity_excel_path): os.remove(self.activity_excel_path)
            if self.summary_excel_path and os.path.exists(self.summary_excel_path): os.remove(self.summary_excel_path)
        except Exception as e: messagebox.showerror("Error", f"Failed to create Zip: {e}")

    # ==========================================
    # STABLE PDF GENERATION (FIX FOR EXE ERROR)
    # ==========================================
    def create_unified_report_stable(self):
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
        except ImportError:
            messagebox.showerror("Error", "ReportLab library missing. Rebuild EXE with reportlab installed.")
            return

        today_str = date.today().strftime("%Y%m%d")
        base_name = f"Unified_Report_{self.username}_{today_str}"
        temp_pdf = os.path.join(self.screenshot_dir, f"{base_name}_temp.pdf")
        final_pdf = os.path.join(os.path.expanduser("~"), "Downloads", f"{base_name}_SECURE.pdf")

        # Build PDF Elements
        elements = []
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        normal_style = styles['Normal']

        # Header
        elements.append(Paragraph(f"Daily Report - {self.username}", title_style))
        elements.append(Paragraph(f"Date: {date.today().strftime('%B %d, %Y')}", normal_style))
        elements.append(Spacer(1, 20))

        # Part 1: Table
        elements.append(Paragraph("Part 1: Activity Log", styles['Heading2']))
        elements.append(Spacer(1, 10))

        # Table Data
        table_data = [['Start', 'End', 'Activity', 'Duration']]
        total_work_seconds = 0
        
        for entry in self.log_data:
            raw = entry['Activity Duration']
            start = entry['Start Time']
            end = entry['End Time']
            activity = raw
            duration = "-"
            
            if " for " in raw:
                parts = raw.split(" for ")
                activity = parts[0]
                duration = parts[1]
                if "Worked" in activity:
                    total_work_seconds += self.parse_duration(raw).total_seconds()
            
            table_data.append([start, end, activity, duration])

        # Table Style
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))

        # Total Calculation Display
        th, tr = divmod(int(total_work_seconds), 3600)
        tm, ts = divmod(tr, 60)
        total_str = f"{th:02d}h {tm:02d}m {ts:02d}s"
        
        summary_data = [["Total Working Time Today", total_str]]
        sum_table = Table(summary_data)
        sum_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (1, 0), (1, 0), colors.green),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(sum_table)
        elements.append(PageBreak())

        # Part 2: Screenshots
        elements.append(Paragraph("Part 2: Screen Captures", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        files = [os.path.join(self.screenshot_dir, f) for f in os.listdir(self.screenshot_dir) 
                 if f.startswith(f"screenshot_{today_str}") and f.endswith(".png")]
        files.sort()

        for i in range(0, len(files), 2):
            # Row of images
            img_row = []
            try:
                img1 = Image(files[i], width=250, height=150) # Resized for PDF
                img_row.append(img1)
                if i+1 < len(files):
                    img2 = Image(files[i+1], width=250, height=150)
                    img_row.append(img2)
            except: pass
            
            if img_row:
                img_table = Table([img_row])
                elements.append(img_table)
                elements.append(Spacer(1, 10))

        # Build PDF
        try:
            doc = SimpleDocTemplate(temp_pdf, pagesize=letter)
            doc.build(elements)
            
            # Encrypt
            reader = PdfReader(temp_pdf)
            writer = PdfWriter()
            for page in reader.pages: writer.add_page(page)
            writer.encrypt(self.report_password)
            
            with open(final_pdf, "wb") as f: writer.write(f)
            os.remove(temp_pdf)
            
            messagebox.showinfo("Success", f"Reports Generated!\n\n1. PDF Report: {final_pdf}\n2. Excel Reports (Zipped): Saved to Downloads folder.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate PDF: {e}")

    # ... [Helpers Unchanged] ...
    def check_internet(self):
        try:
            import socket
            socket.create_connection(("8.8.8.8", 53), timeout=3)
            return True
        except: return False

    def check_lock(self):
        try: return ctypes.windll.User32.GetForegroundWindow() == 0
        except: return False

    def parse_duration(self, d_str):
        try:
            if " for " in d_str:
                parts = d_str.split(" for ")[1].strip().split(' ')
                h = int(parts[0].replace('h',''))
                m = int(parts[1].replace('m',''))
                s = int(parts[2].replace('s',''))
                return timedelta(hours=h, minutes=m, seconds=s)
        except: pass
        return timedelta(0)

    def load_daily_backup(self):
        if os.path.exists(self.daily_backup_file):
            try:
                with open(self.daily_backup_file, 'r') as f:
                    self.log_data = json.load(f)
            except: self.log_data = []

    def save_backup(self):
        try:
            with open(self.daily_backup_file, 'w') as f:
                json.dump(self.log_data, f)
        except: pass

class MainApplication:
    def __init__(self, root):
        self.root = root
        self.root.title("Employee Monitoring Suite")
        self.root.geometry("550x650") 
        self.data_dir = os.path.join(os.path.expanduser("~"), "WorkLogData")
        os.makedirs(self.data_dir, exist_ok=True)
        self.config_file = os.path.join(self.data_dir, "config.txt")
        self.username = self.load_username()
        if not self.username: self.show_login_frame()
        else: self.show_main_interface()

    def load_username(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, 'r') as f: return f.read().strip()
        return None

    def save_username(self, name):
        with open(self.config_file, 'w') as f: f.write(name)
        self.username = name

    def show_login_frame(self):
        self.login_frame = tk.Frame(self.root)
        self.login_frame.pack(expand=True, fill='both', padx=20, pady=20)
        tk.Label(self.login_frame, text="Welcome Employee", font=("Arial", 16, "bold")).pack(pady=20)
        tk.Label(self.login_frame, text="Please enter your name:", font=("Arial", 12)).pack()
        self.name_entry = tk.Entry(self.login_frame, font=("Arial", 12))
        self.name_entry.pack(pady=10)
        tk.Button(self.login_frame, text="Start Session", command=self.submit_login, bg="#2196F3", fg="white", font=("Arial", 12, "bold")).pack(pady=20)

    def submit_login(self):
        name = self.name_entry.get().strip()
        if name:
            self.save_username(name)
            self.login_frame.destroy()
            self.show_main_interface()
        else: messagebox.showwarning("Error", "Name is required")

    def show_main_interface(self):
        self.app = UnifiedMonitorApp(self.root, self.username, self.data_dir)

if __name__ == "__main__":
    root = tk.Tk()
    try: ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except: pass
    app = MainApplication(root)
    root.mainloop()