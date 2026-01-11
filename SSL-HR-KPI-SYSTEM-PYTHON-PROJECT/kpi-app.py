import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import openpyxl
import os
import datetime
import sys

# ==========================================
# CONFIGURATION
# ==========================================
# Columns mapping
REQUIRED_COLUMNS = {
    "Production": [4, 5, 6],           # E, F, G
    "Planning": [7, 8],                # H, I
    "Yarn": [9, 10, 11],               # J, K, L
    "Store": [12, 13, 14],             # M, N, O
    "Commercial Dept.": [15, 16, 17, 18], # P, Q, R, S
    "ERP": [19, 20],                   # T, U
    "HR": [21, 22],                    # V, W
    "Reporting Boss": [23]             # X
}

# Excel Configuration
DATA_START_ROW = 14  # Row where employee data starts
ID_COL = 2           # Column B contains the ID (1=A, 2=B)

# Global variable to store the path of the Master File
current_master_path = ""

def get_grade(score):
    """Calculates grade based on score."""
    if score >= 90: return 'A'
    elif score >= 80: return 'B'
    elif score >= 60: return 'C'
    else: return 'F'

def auto_detect_master():
    """Tries to find the Master file in the current folder automatically."""
    global current_master_path
    # Look for files containing 'Master-File' (case insensitive)
    files = [f for f in os.listdir('.') if f.endswith('.xlsx') and "master-file" in f.lower()]
    
    if files:
        # Pick the first one found
        current_master_path = os.path.abspath(files[0])
        lbl_master_status.config(text=f"Ready: {os.path.basename(current_master_path)}", fg="green")
        btn_change_master.config(text="Change Master File")
    else:
        current_master_path = ""
        lbl_master_status.config(text="Not Found. Please select Master File manually.", fg="red")

def select_master_manually():
    """Allows user to manually pick the Master File."""
    global current_master_path
    path = filedialog.askopenfilename(title="Select the MASTER KPI File", filetypes=[("Excel Files", "*.xlsx")])
    if path:
        current_master_path = path
        lbl_master_status.config(text=f"Ready: {os.path.basename(path)}", fg="green")

def find_kpi_sheet_in_dept_file(filepath):
    """Reads a Dept file and finds the sheet with the actual data."""
    try:
        xls = pd.ExcelFile(filepath)
        for sheet in xls.sheet_names:
            # Read header area
            df = pd.read_excel(xls, sheet_name=sheet, header=None, nrows=20)
            df_str = df.astype(str)
            # Check for keywords
            if df_str.apply(lambda x: x.str.contains("Merchandiser's Name|ID No", regex=True)).any().any():
                return pd.read_excel(xls, sheet_name=sheet, header=None)
        # Fallback to first sheet
        return pd.read_excel(xls, sheet_name=0, header=None)
    except Exception as e:
        return None

def manage_monthly_sheet(wb):
    """Handles creating the new month's sheet if it doesn't exist."""
    current_date = datetime.datetime.now()
    # E.g., "KPI December-2025"
    new_sheet_name = f"KPI {current_date.strftime('%B-%Y')}"
    
    # 1. Check if sheet already exists
    if new_sheet_name in wb.sheetnames:
        return new_sheet_name, False # False means we didn't just create it

    # 2. Find a template sheet to copy (Look for any sheet with 'KPI' in name)
    kpi_sheets = [s for s in wb.sheetnames if "KPI" in s]
    if not kpi_sheets:
        raise Exception("No 'KPI' sheet found in Master File to copy format from.")

    # Copy the last available KPI sheet
    source_sheet = wb[kpi_sheets[-1]]
    target_sheet = wb.copy_worksheet(source_sheet)
    target_sheet.title = new_sheet_name
    
    # 3. Clear old data from the new sheet (Keep headers/IDs, remove scores)
    max_row = target_sheet.max_row
    for r in range(DATA_START_ROW, max_row + 1):
        # Clear Data Columns (E to X -> 5 to 24)
        for c in range(5, 25): 
            target_sheet.cell(row=r, column=c).value = None
        # Clear Formula/Result Columns (Y to AA -> 25 to 27)
        target_sheet.cell(row=r, column=25).value = None
        target_sheet.cell(row=r, column=26).value = None
        target_sheet.cell(row=r, column=27).value = None
        
    return new_sheet_name, True # True means we created a new one

def run_update_process():
    """Main execution logic."""
    global current_master_path
    
    # 1. Validate Master File
    if not current_master_path or not os.path.exists(current_master_path):
        messagebox.showerror("Error", "Please select a valid Master File first!")
        return

    # 2. Select Department Files
    dept_files = filedialog.askopenfilenames(title="Select Department KPI Files", filetypes=[("Excel Files", "*.xlsx")])
    if not dept_files:
        return # User cancelled

    # Check: Did user accidentally select the master file as an input?
    for d_file in dept_files:
        if os.path.abspath(d_file) == os.path.abspath(current_master_path):
            messagebox.showerror("Error", f"You selected the Master File as an input!\n\nSkip: {os.path.basename(d_file)}\n\nPlease select only Department files.")
            return

    # 3. Start Processing
    log_box.delete(1.0, tk.END)
    log_box.insert(tk.END, f"Target Master File: {os.path.basename(current_master_path)}\n")
    log_box.insert(tk.END, "-"*40 + "\n")
    root.update()

    try:
        # Load Master Workbook
        wb = openpyxl.load_workbook(current_master_path)
        
        # Prepare Monthly Sheet
        active_sheet_name, created_new = manage_monthly_sheet(wb)
        ws = wb[active_sheet_name]
        
        if created_new:
            log_box.insert(tk.END, f"[NEW MONTH] Created sheet: {active_sheet_name}\n")
        else:
            log_box.insert(tk.END, f"[UPDATE] Updating sheet: {active_sheet_name}\n")
        
        total_updates = 0

        # Loop through Dept Files
        for d_file in dept_files:
            fname = os.path.basename(d_file)
            log_box.insert(tk.END, f"\nReading: {fname}...")
            root.update()
            
            # Read Dept Data
            df_dept = find_kpi_sheet_in_dept_file(d_file)
            if df_dept is None:
                log_box.insert(tk.END, " -> FAIL (Not a valid Excel)\n")
                continue

            # Identify Active Columns (Which Dept is this?)
            active_cols = []
            # Safety: Ensure file has enough rows
            if df_dept.shape[0] > DATA_START_ROW:
                df_slice = df_dept.iloc[DATA_START_ROW-1:, :]
                
                for dept, cols in REQUIRED_COLUMNS.items():
                    # Ensure columns exist in this file
                    valid_cols = [c for c in cols if c < df_dept.shape[1]]
                    if not valid_cols: continue
                    
                    # Check for data
                    subset = df_slice.iloc[:, valid_cols].apply(pd.to_numeric, errors='coerce')
                    if subset.notna().any().any() and subset.sum().sum() > 0:
                        log_box.insert(tk.END, f" -> Detected: {dept}")
                        active_cols.extend(valid_cols)
            
            if not active_cols:
                log_box.insert(tk.END, " -> No data found.\n")
                continue

            # Map Data: ID -> {Col: Value}
            # Note: IDs are in Column B (index 1) in the dataframe
            update_map = {}
            for idx, row in df_dept.iloc[DATA_START_ROW-1:].iterrows():
                # Safety check on column bounds
                if df_dept.shape[1] <= 1: continue
                
                merch_id = str(row[1]).strip()
                if pd.isna(merch_id) or merch_id == 'nan': continue
                
                row_data = {}
                for c in active_cols:
                    val = row[c]
                    if pd.notna(val):
                        row_data[c] = val
                if row_data:
                    update_map[merch_id] = row_data

            # Apply to Master Sheet
            file_updates = 0
            for r in range(DATA_START_ROW, ws.max_row + 1):
                cell_id = ws.cell(row=r, column=ID_COL).value
                if cell_id:
                    s_id = str(cell_id).strip()
                    if s_id in update_map:
                        for c_idx, val in update_map[s_id].items():
                            ws.cell(row=r, column=c_idx + 1).value = val
                        file_updates += 1
            
            log_box.insert(tk.END, f"\n   -> Merged {file_updates} rows.")
            total_updates += 1

        # 4. Final Recalculation
        log_box.insert(tk.END, "\n\nRecalculating Totals & Grades... ")
        root.update()
        
        for r in range(DATA_START_ROW, ws.max_row + 1):
            total = 0
            # Sum Score Columns (E to X -> 5 to 24)
            for c in range(5, 25):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, (int, float)):
                    total += val
            
            # Write Totals
            ws.cell(row=r, column=25).value = total      
            ws.cell(row=r, column=26).value = total/100  
            ws.cell(row=r, column=26).number_format = '0%'
            ws.cell(row=r, column=27).value = get_grade(total)

        # 5. Save
        try:
            wb.save(current_master_path)
            log_box.insert(tk.END, "Done!\n")
            log_box.insert(tk.END, "-"*40 + "\n")
            log_box.insert(tk.END, "SUCCESS! Master File Saved.\n")
            messagebox.showinfo("Success", "Master File Updated Successfully!")
        except PermissionError:
            messagebox.showerror("Error", "Could not save Master File.\nIt is currently OPEN.\n\nPlease close Excel and try again.")
            log_box.insert(tk.END, "\nERROR: File is open. Save Failed.")

    except Exception as e:
        log_box.insert(tk.END, f"\nCRITICAL ERROR: {str(e)}")
        messagebox.showerror("Error", str(e))

# ==========================================
# GUI SETUP
# ==========================================
root = tk.Tk()
root.title("KPI Automation Tool v2.0")
# Increased height to accommodate the new footer
root.geometry("600x750")
root.configure(bg="#f0f0f0")

# --- Header ---
header_frame = tk.Frame(root, bg="#333", pady=10)
header_frame.pack(fill="x")
tk.Label(header_frame, text="Merchandiser KPI Automation", fg="white", bg="#333", font=("Arial", 14, "bold")).pack()

# --- Section 1: Master File ---
frame_master = tk.LabelFrame(root, text="1. Master File Setup", font=("Arial", 10, "bold"), bg="#f0f0f0", padx=10, pady=10)
frame_master.pack(fill="x", padx=15, pady=10)

lbl_master_status = tk.Label(frame_master, text="Searching...", font=("Arial", 10), bg="#f0f0f0")
lbl_master_status.pack(side="left", fill="x", expand=True)

btn_change_master = tk.Button(frame_master, text="Select Master File", command=select_master_manually)
btn_change_master.pack(side="right")

# --- Section 2: Action ---
frame_action = tk.LabelFrame(root, text="2. Update Data", font=("Arial", 10, "bold"), bg="#f0f0f0", padx=10, pady=10)
frame_action.pack(fill="x", padx=15, pady=5)

lbl_instr = tk.Label(frame_action, text="Click below to upload Dept files (Production, Yarn, etc.)\nThe code will merge them into the Master File above.", bg="#f0f0f0", justify="left")
lbl_instr.pack(pady=5)

btn_run = tk.Button(frame_action, text="Upload Department Files & Run", command=run_update_process, bg="#008CBA", fg="white", font=("Arial", 12, "bold"), height=2)
btn_run.pack(fill="x", padx=10, pady=5)

# --- Log ---
frame_log = tk.Frame(root, padx=15, pady=10, bg="#f0f0f0")
frame_log.pack(fill="both", expand=True)
tk.Label(frame_log, text="Process Log:", bg="#f0f0f0", font=("Arial", 9, "bold")).pack(anchor="w")

scrollbar = tk.Scrollbar(frame_log)
scrollbar.pack(side="right", fill="y")

log_box = tk.Text(frame_log, height=10, yscrollcommand=scrollbar.set, font=("Consolas", 9))
log_box.pack(fill="both", expand=True)
scrollbar.config(command=log_box.yview)

# --- DEVELOPER INFO SECTION (NEW) ---
frame_dev = tk.Frame(root, bg="#e8e8e8", padx=10, pady=10, relief=tk.RIDGE, bd=2)
frame_dev.pack(fill="x", padx=15, pady=15)

# Developer Branding
tk.Label(frame_dev, text="This app is In house Developed By", bg="#e8e8e8", font=("Arial", 10)).pack(pady=(0, 2))
tk.Label(frame_dev, text="Sonia and Sweaters Limited", bg="#e8e8e8", fg="#C0392B", font=("Helvetica", 14, "bold")).pack(pady=(0, 10))

# Contact Information
tk.Label(frame_dev, text="For any types of inconvience please Contact with:", bg="#e8e8e8", font=("Arial", 9, "bold", "underline")).pack(pady=(5, 2))
tk.Label(frame_dev, text="Prottoy Saha", bg="#e8e8e8", font=("Arial", 9, "bold")).pack()
tk.Label(frame_dev, text="Automation Engineer", bg="#e8e8e8", font=("Arial", 9)).pack()
tk.Label(frame_dev, text="Sonia and Sweaters Limited", bg="#e8e8e8", font=("Arial", 9)).pack()
tk.Label(frame_dev, text="+8801745547578", bg="#e8e8e8", font=("Arial", 9)).pack()
tk.Label(frame_dev, text="prottoy.saha@soniagroup.com", bg="#e8e8e8", font=("Arial", 9, "italic")).pack()

# --- Auto-Run Setup ---
root.after(500, auto_detect_master) # Auto-detect master file on load

root.mainloop()